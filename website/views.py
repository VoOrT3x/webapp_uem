from flask import Blueprint, render_template,request
import matplotlib
matplotlib.use('agg')



views = Blueprint("views", __name__)

@views.route("/")

def home():
    return render_template("base.html")


@views.route("/map")
def map():
    import geopandas as gpd
    import json

    # Folder with Shapefiles
    folder_path = r'\website\static'

    # Exact shape file
    shapefile_path = folder_path + "\Distritos.shp"

    # Load shapefile dos distritos
    districts = gpd.read_file(shapefile_path)

    # Convert districts to GeoJSON
    districts_json = districts.to_crs(epsg='4326').to_json()

    return render_template("map.html", districts_json=districts_json)



@views.route("/map/district/<district_name>")
def district_page(district_name):
    import pulp 
    import openpyxl
    from flask import jsonify

    
    book= openpyxl.load_workbook(r"website\static\Dados_py.xlsx")
    worksheet=book["Dados"]
    
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if row[0]==district_name:
            Irr=str(row[1])
            Temp=str(row[2])
            Vent=str(row[3])

    Irr=float(Irr)
    Temp=float(Temp)
    Vent=float(Vent)
    
    return render_template("district.html", district_name=district_name,Temp=Temp,Irr=Irr,Vent=Vent)


@views.route("/map/district/<district_name>/calculate", methods=["POST"])
def calculate(district_name):
    import pulp 
    import openpyxl
    from flask import jsonify
    import matplotlib.pyplot as plt
    import numpy as np
    import mpl_toolkits.mplot3d as Axes3D
    import math
    from matplotlib import cm
    from matplotlib.ticker import LinearLocator
    demand = float(request.form["demand"])
    PF = float(request.form["pf"])
    PE= float(request.form["pe"])
    book = openpyxl.load_workbook(r"website\static\Dados_py.xlsx")
    worksheet = book["Dados"]
    
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if row[0] == district_name:
            G = float(row[1])
            T = float(row[2])
            V = float(row[3])
            Q = float(row[4])
    
    
    PM = float(PF*G/1000*(1-0.0006*(T-25)))
    PT = round(0.9561*math.exp(-((V-25.5731)/(8.2016))**2)+0.5874*math.exp(-((V-11.0979)/(4.1853))**2)+0.6075*math.exp(-((V-16.3506)/(5.4920))**2),4)
    print (PM)
    print(PT)
    model = pulp.LpProblem(name="Lp", sense=pulp.const.LpMinimize)

    x_1 = pulp.LpVariable(name="Módulos fotovoltaicos", lowBound=0)
    x_2 = pulp.LpVariable(name="Turbinas eólicas", lowBound=0)
    x_3 = pulp.LpVariable(name="Altura do subsistema hídrico", lowBound=0)
    
    obj_func =4647*x_1*PF+8802*x_2*PE+8545*Q*x_3
    
    model += obj_func
    
    model += (8670*x_1*PM+8670*x_2*PT*PE+17218.7*Q*x_3 >= demand*10**6)
    model += (8670*x_2*PT*PE<= 0.3*demand*10**6)
    model += (8670*x_1*PM<= 0.80*demand*10**6)
    model += (x_3<= 25)
    model += (17218.7*Q*x_3<= 0.1*demand*10**6)
    model += (17218.7*Q*x_3<=8670*x_2*PT*PE )

    status = model.solve()
    
    results_dict = {}
    results_dict["Módulos_fotovoltaicos"]= math.ceil(x_1.varValue)
    results_dict["Turbinas_eólicas"]= math.ceil(x_2.varValue)
    results_dict["Altura_do_subsistema_hídrico"]= round(x_3.varValue,1)
    results_dict["Custo"]= round(4647*results_dict["Módulos_fotovoltaicos"]*PF+8802*results_dict["Turbinas_eólicas"]*PE+8545*Q*results_dict["Altura_do_subsistema_hídrico"],2)
    results_dict["PMP"]= round(PM,2)
    results_dict["PV"]= round(PT,2)
    
    fig = plt.figure(figsize=(12, 5))
    ax1 = fig.add_subplot(1,2,1, projection='3d')
    x1 = np.arange(0, 80, 5)
    y1 = np.arange(0, 2000, 100)
    x2 = np.arange(0,25,1)
    y2= 0.9561*np.exp(-((x2-25.5731)/(8.2016))**2)+0.5874*np.exp(-((x2-11.0979)/(4.1853))**2)+0.6075*np.exp(-((x2-16.3506)/(5.4920))**2)
    
    X1, Y1 = np.meshgrid(x1, y1)
    Z1 = ((PF*Y1/1000)*(1-0.0006*(X1-25)))
    # Plotting the feasible region    
    ax1.plot_surface(X1, Y1, Z1, alpha=1,cmap='viridis_r',
                       linewidth=0, antialiased=False,edgecolor = 'k',label=f'Demanda Energetica')
    ax1.set_zlim(0, round(PM,2)+0.5)
    ax1.set_xlabel("Temperatura")
    ax1.set_ylabel("Irradiação solar")
    ax1.zaxis.set_major_locator(LinearLocator(10))
    ax1.set_title("Variação da potência máxima do módulos fotovoltaicos", fontweight="bold")
    
    ax2 = fig.add_subplot(1, 2, 2)
    ax2.set_title("Variação do coeficiente de potência das turbinas eólicas",fontweight="bold")
    ax2.plot(x2, y2, label='y = 2x + 1', color='blue') 
    ax2.grid(True)  
    plt.subplots_adjust(wspace=0.6,bottom=0.1)

    # Save the plot image to a file
    plot_image = 'D:\\Trabalho de Licenciatura\\New folder\\Python\\WebApp\\website\\static\\images\\plot.png'
    plt.savefig(plot_image)

    # Close the figure to free up resources
    plt.close()

    # Pass the image path as a variable to the template
    return jsonify(results_dict=results_dict, plot_image=plot_image)
